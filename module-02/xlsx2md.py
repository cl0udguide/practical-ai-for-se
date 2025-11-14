#!/usr/bin/env python3
"""
Simple CLI tool to convert Excel spreadsheets to Markdown with formula evaluation.
This script uses openpyxl to replace formulas with their calculated values before
conversion, addressing docling's limitation of not evaluating formula results.
"""

import sys
import tempfile
from pathlib import Path

import openpyxl
from docling.datamodel.base_models import InputFormat
from docling.document_converter import DocumentConverter


def replace_formulas_with_values(input_file: str, output_file: str):
    """
    Replace all formulas in an Excel file with their calculated values.

    Args:
        input_file: Path to the input Excel file
        output_file: Path to the output Excel file with formulas replaced
    """
    print(f"  - Loading workbook and evaluating formulas...")

    # Load the workbook with data_only=False to read formulas
    wb = openpyxl.load_workbook(input_file, data_only=False)

    # Load again with data_only=True to get calculated values
    wb_values = openpyxl.load_workbook(input_file, data_only=True)

    formula_count = 0
    for sheetname in wb.sheetnames:
        ws_formula = wb[sheetname]
        ws_values = wb_values[sheetname]

        # Iterate through all cells and replace formula with value
        for row in ws_formula.iter_rows():
            for cell in row:
                if cell.data_type == "f":  # Cell contains a formula
                    cell.value = ws_values[cell.coordinate].value
                    formula_count += 1

    # Save to new file
    wb.save(output_file)
    wb.close()
    wb_values.close()

    print(f"  - Replaced {formula_count} formula(s) with calculated values")


def convert_xlsx_to_markdown(input_file: str, output_file: str = None):
    """
    Convert an Excel file to Markdown format with formula evaluation.

    Args:
        input_file: Path to the input Excel file
        output_file: Path to the output Markdown file (optional)
    """
    input_path = Path(input_file)

    # Validate input file
    if not input_path.exists():
        print(f"Error: File '{input_file}' not found.")
        sys.exit(1)

    if input_path.suffix.lower() not in [".xlsx", ".xlsm"]:
        print(f"Error: File must be a .xlsx or .xlsm file, got '{input_path.suffix}'")
        sys.exit(1)

    # Set output path
    if output_file is None:
        output_path = input_path.with_suffix(".md")
    else:
        output_path = Path(output_file)

    print(f"Converting '{input_file}' to Markdown...")

    # Create a temporary file for the formula-evaluated version
    temp_file = None
    try:
        # Create temporary file with formulas replaced
        with tempfile.NamedTemporaryFile(mode="w", suffix=".xlsx", delete=False) as tmp:
            temp_file = tmp.name

        # Replace formulas with their calculated values
        replace_formulas_with_values(str(input_path), temp_file)

        print(f"  - Converting to Markdown...")

        # Configure converter for Excel files
        converter = DocumentConverter(allowed_formats=[InputFormat.XLSX])

        # Convert the temporary document
        result = converter.convert(temp_file)

        # Export to Markdown
        markdown_content = result.document.export_to_markdown()

        # Write to output file
        output_path.write_text(markdown_content, encoding="utf-8")

        print(f"✓ Successfully converted to '{output_path}'")
        print(f"  - Sheets converted: {len(result.document.pages)}")

    except Exception as e:
        print(f"Error during conversion: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)
    finally:
        # Clean up temporary file
        if temp_file and Path(temp_file).exists():
            Path(temp_file).unlink()


def print_usage():
    """Print usage information."""
    print(
        """
Usage: python xlsx2md.py <input.xlsx> [output.md]

Arguments:
  input.xlsx    Path to the Excel file to convert (.xlsx or .xlsm)
  output.md     (Optional) Path to the output Markdown file
                If not specified, uses the same name as input with .md extension

Features:
  - Automatically evaluates all formulas before conversion
  - Converts all sheets in the workbook
  - Preserves table structure in Markdown format

Examples:
  python xlsx2md.py spreadsheet.xlsx
  python xlsx2md.py spreadsheet.xlsx output.md
  python xlsx2md.py "Financial Report.xlsx" report.md

Requirements:
  - openpyxl library (install with: pip install openpyxl)
  - Excel file must have been saved with calculated formula values
"""
    )


def main():
    """Main entry point for the CLI."""
    args = sys.argv[1:]

    # Check for help flag
    if not args or "--help" in args or "-h" in args:
        print_usage()
        sys.exit(0)

    # Parse arguments
    input_file = None
    output_file = None

    for arg in args:
        if input_file is None:
            input_file = arg
        elif output_file is None and not arg.startswith("--"):
            output_file = arg

    if input_file is None:
        print("Error: No input file specified.")
        print_usage()
        sys.exit(1)

    # Convert the file
    convert_xlsx_to_markdown(input_file, output_file)


if __name__ == "__main__":
    main()
